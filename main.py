import re
from datetime import datetime

import pandas as pd
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from functions.calculate_coordinates import calculate_coordinates
from functions.get_bnr_exchange_rate import get_bnr_exchange_rate
from functions.get_country_code_from_address import get_country_code_from_address
from functions.get_delivery_location import get_delivery_location


class Constants:
    # Proportions for invoice sections
    PROPORTIONS = {"section_1": (0.0, 0.0, 1.0, 0.16), "section_2": (0.0, 0.16, 0.46, 0.54),
                   "section_3": (0.46, 0.16, 1.0, 0.54), "section_4": (0.0, 0.54, 1.0, 0.93),
                   "section_5": (0.0, 0.93, 1.0, 1.0), }

    FINAL_HEADERS = ["Nr. Crt.", "Firma", "Nr. Factura marfa", "Cod NC8", "origine", "destinatie", "Val. Fact. EURO",
                     "Greutate neta", "data expeditiei", "Curs valutar", "Valoare RON", "Vat/cumparator", "LOC LIVRARE",
                     "Conditie livrare", "%", "transport", "STATistica"]

    COLUMN_MAPPING = {"company": "Firma", "invoice_number": "Nr. Factura marfa", "nc8_code": "Cod NC8",
                      "origin": "origine", "destination": "destinatie", "invoice_value_eur": "Val. Fact. EURO",
                      "net_weight": "Greutate neta", "shipment_date": "data expeditiei",
                      "exchange_rate": "Curs valutar", "value_ron": "Valoare RON", "vat_number": "Vat/cumparator",
                      "delivery_location": "LOC LIVRARE", "delivery_condition": "Conditie livrare"}


class InvoiceProcessor:
    def __init__(self, paths):
        self.input_paths = paths
        self.df = pd.DataFrame(
            columns=["company", "invoice_number", "nc8_code", "origin", "destination", "invoice_value_eur",
                     "net_weight", "shipment_date", "exchange_rate", "value_ron", "vat_number", "delivery_location",
                     "delivery_condition"])

    def process_invoices(self):
        for idx, pdf_path in enumerate(self.input_paths):
            with pdfplumber.open(pdf_path) as pdf:
                first_page = pdf.pages[0]
                page_width, page_height = first_page.width, first_page.height

                # self._extract_sections_as_images(first_page, page_width, page_height)

                section_2_text = self._extract_section_text(first_page, "section_2", page_width, page_height)
                company = self._extract_field(section_2_text, r"Our payment address\n(.+)", "Unknown")
                print("Company:", company)

                section_1_text = self._extract_section_text(first_page, "section_1", page_width, page_height)
                invoice_number = self._extract_invoice_number(section_1_text)
                print("Invoice Number:", invoice_number)

                nc8_codes = self._extract_nc8_codes(pdf)
                print("NC8 Codes:", nc8_codes)

                origin = get_country_code_from_address(
                    self._extract_field(section_2_text, r"Our payment address\n(.+?)\nPayment date", "Unknown",
                                        re.DOTALL))
                print("Origin:", origin)

                section_3_text = self._extract_section_text(first_page, "section_3", page_width, page_height)
                destination = get_country_code_from_address(
                    self._extract_field(section_3_text, r"Invoiced to : (.+?)\nCredit transfer", "Unknown", re.DOTALL))
                print("Destination:", destination)

                invoice_value_eur, invoice_value_ron = self._extract_invoice_values(pdf.pages[-1], page_width,
                                                                                    page_height)
                print("Invoice Value EUR:", invoice_value_eur)
                print("Invoice Value RON:", invoice_value_ron)

                net_weight = self._extract_net_weight(pdf.pages[-1])
                print("Net Weight:", net_weight)

                shipment_date = self._extract_field(pdf.pages[-1].extract_text(),
                                                    r"Transportation date: (\d{2}\.\d{2}\.\d{4})", None)
                shipment_date = datetime.strptime(shipment_date, "%d.%m.%Y") if shipment_date else datetime.now()
                print("Shipment Date:", shipment_date.strftime("%d.%m.%Y"))

                exchange_rate = get_bnr_exchange_rate(shipment_date, "RON")
                print("Exchange Rate:", exchange_rate)

                delivery_location = get_delivery_location(section_1_text)
                print("Delivery Location:", delivery_location)

                vat_number = self._extract_field(section_3_text, r"Tax number : (\w+)", "Unknown")
                print("VAT Number:", vat_number)

                delivery_condition = self._extract_field(section_2_text, r"Incoterms : (\w+)", "Unknown")
                print("Delivery Condition:", delivery_condition)

                self.df.loc[idx] = [company, invoice_number, ", ".join(nc8_codes), origin, destination,
                                    invoice_value_eur, net_weight, shipment_date.strftime("%d.%m.%Y"), exchange_rate,
                                    invoice_value_ron, vat_number, delivery_location, delivery_condition]

    @staticmethod
    def _extract_sections_as_images(page, page_width, page_height):
        for section_name, proportion in Constants.PROPORTIONS.items():
            section_coords = calculate_coordinates(page_width, page_height, proportion)
            section_image = page.within_bbox(section_coords).to_image()
            image_path = f"output/{section_name}.png"
            section_image.save(image_path)

    @staticmethod
    def _extract_section_text(page, section_name, page_width, page_height):
        section_coords = calculate_coordinates(page_width, page_height, Constants.PROPORTIONS[section_name])
        return page.within_bbox(section_coords).extract_text()

    @staticmethod
    def _extract_field(text, pattern, default, flags=0):
        match = re.search(pattern, text, flags)
        return match.group(1).strip() if match else default

    @staticmethod
    def _extract_invoice_number(section_1_text):
        lines = section_1_text.split("\n")
        return lines[-1].split(" ")[0] if lines else "Unknown"

    @staticmethod
    def _extract_nc8_codes(pdf):
        codes = []
        for page in pdf.pages:
            codes += re.findall(r"Commodity Code : (\d+)", page.extract_text())
        return codes

    @staticmethod
    def _extract_invoice_values(last_page, page_width, page_height):
        bbox_ratio = (1125 / 1653, 2054 / 2339, 1552 / 1653, 2182 / 2339)
        bbox_coords = calculate_coordinates(page_width, page_height, bbox_ratio)
        bbox_text = last_page.within_bbox(bbox_coords).extract_text()
        lines = bbox_text.splitlines() if bbox_text else []
        last_line = lines[-1] if lines else ""
        if "*" in last_line:
            currency, currency_value = last_line.split("*")[0].strip().lower(), last_line.split("*")[-1].strip()
            if currency == "eur":
                return float(currency_value.replace(",", "").replace(".", ".")), 0
            elif currency == "ron":
                return 0, float(currency_value.replace(".", "").replace(",", "."))
        return 0, 0

    @staticmethod
    def _extract_net_weight(last_page):
        match = re.search(r"Net weight\s+(.+?)\s+KG", last_page.extract_text())
        return float(match.group(1).replace(",", ".").replace(".", "")) if match else 0


def add_totals(df, columns_to_total, group_by=None):
    """
    Adaugă rânduri de totaluri pentru coloanele specificate folosind funcția `SUM` din Excel.
    Totalurile sunt adăugate imediat după fiecare grup (dacă `group_by` este specificat).
    """
    total_rows = []
    current_row_offset = 1  # Începem de la rândul 1 după antet

    if group_by:
        grouped = df.groupby(group_by, sort=False)
        results = []
        for name, group in grouped:
            group = group.reset_index(drop=True)
            results.append(group)  # Adăugăm grupul original în listă

            # Adăugăm rândul de totaluri
            total_row = {col: "" for col in df.columns}
            total_row[group_by] = f"Total {name}"
            total_row["Nr. Crt."] = ""

            # Adaugăm formulele de totaluri pentru fiecare coloană specificată
            for col in columns_to_total:
                # Determinăm litera coloanei în funcție de indexul său
                col_letter = chr(65 + df.columns.get_loc(col))  # Convertim indexul în literă de coloană Excel
                start_row = current_row_offset + 1  # Prima linie a grupului în Excel
                end_row = start_row + len(group) - 1  # Ultima linie a grupului în Excel
                total_row[col] = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"

            # Actualizăm offset-ul curent
            current_row_offset += len(group) + 1  # Adăugăm numărul de rânduri din grup + 1 pentru total

            # Adăugăm rândul de totaluri în lista rezultat
            results.append(pd.DataFrame([total_row]))

        # Combinăm toate grupurile și rândurile de totaluri
        df = pd.concat(results, ignore_index=True)
    else:
        # Total general dacă nu există grupare
        total_row = {col: "" for col in df.columns}
        total_row["Nr. Crt."] = "Total"

        # Adaugăm formulele de totaluri pentru coloanele specificate
        for col in columns_to_total:
            col_letter = chr(65 + df.columns.get_loc(col))  # Convertim indexul în literă de coloană Excel
            start_row = 2  # Prima linie de date (Excel începe de la rândul 2 după antet)
            end_row = len(df) + 1  # Ultimul rând de date
            total_row[col] = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"

        df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

    return df


def add_excel_formulas(df):
    """
    Adaugă formule pentru coloanele `transport` și `STATistica` folosind referințele celulelor din Excel.
    """
    for i, row in df.iterrows():
        excel_row = i + 2  # Rândul în Excel (începe de la 2 deoarece 1 este pentru antet)

        # Referințele coloanelor în funcție de poziția lor în antet
        cell_curs_valutar = f"J{excel_row}"  # Coloana "Curs valutar"
        cell_greutate_neta = f"H{excel_row}"  # Coloana "Greutate neta"
        cell_valoare_eur = f"G{excel_row}"  # Coloana "Val. Fact. EURO"
        cell_valoare_ron = f"K{excel_row}"  # Coloana "Valoare RON"
        cell_procent = f"O{excel_row}"  # Coloana "%"

        if pd.isna(row["Nr. Crt."]) or str(row["Nr. Crt."]).strip() == "":
            continue

        # Adăugăm formule pentru completarea valorilor lipsă
        if row["Valoare RON"] == 0:
            df.at[i, "Valoare RON"] = f"={cell_valoare_eur}*{cell_curs_valutar}"
        if row["Val. Fact. EURO"] == 0:
            df.at[i, "Val. Fact. EURO"] = f"={cell_valoare_ron}/{cell_curs_valutar}"

        # Formula pentru Transport
        if pd.notna(row["Greutate neta"]) and pd.notna(row["Curs valutar"]):
            df.at[i, "transport"] = f"=28000*{cell_curs_valutar}/147000*{cell_greutate_neta}"
        else:
            df.at[i, "transport"] = ""

        # Formula pentru STATistica
        if pd.notna(row["Valoare RON"]):
            df.at[i, "STATistica"] = f"=ROUND({cell_valoare_ron}+{cell_procent}*{cell_curs_valutar}, 0)"
        else:
            df.at[i, "STATistica"] = ""

    return df


# Funcția pentru extinderea `Cod NC8` pe rânduri separate
def expand_nc8(df):
    """
    Extinde coloana `Cod NC8` astfel încât fiecare cod să apară pe un rând separat.
    """
    rows = []
    for _, row in df.iterrows():
        nc8_codes = row["Cod NC8"].split(", ")
        for i, code in enumerate(nc8_codes):
            new_row = row.copy()
            if i == 0:
                new_row["Nr. Crt."] = len(rows) + 1
            else:
                new_row["Nr. Crt."] = ""
                for col in ["Firma", "Nr. Factura marfa", "origine", "destinatie", "Val. Fact. EURO", "Greutate neta",
                            "data expeditiei", "Curs valutar", "Valoare RON", "Vat/cumparator", "LOC LIVRARE",
                            "Conditie livrare"]:
                    new_row[col] = ""
            new_row["Cod NC8"] = code
            rows.append(new_row)
    return pd.DataFrame(rows)


def add_row_numbers(df):
    """
    Adaugă numere de ordine în coloana `Nr. Crt.`.
    """
    df["Nr. Crt."] = range(1, len(df) + 1)
    return df


def merge_nc8_codes(df):
    """
    Combină codurile NC8 într-o singură celulă per factură, separate prin virgule, eliminând duplicatele.
    """
    for index, row in df.iterrows():
        unique_nc8_codes = sorted(set(row["Cod NC8"].split(", ")))  # Elimină duplicatele și sortează
        df.at[index, "Cod NC8"] = ", ".join(unique_nc8_codes)  # Reunește codurile într-un string
    return df


def generate_excel(data, output_path="processed_invoices.xlsx"):
    """
    Generează un fișier Excel din DataFrame cu toate cerințele specificate.
    """
    # Adaugăm numere de ordine
    data = add_row_numbers(data)

    # Asigurăm ordinea corectă a coloanelor
    data = data.loc[:, [h for h in Constants.FINAL_HEADERS if h not in ['%', 'transport', 'STATistica']]]

    # Combinăm codurile NC8 în celule unice
    data = merge_nc8_codes(data)

    # Grupare pe `Firma` și sortare după `data expeditiei`
    data = data.sort_values(by=["Firma", "data expeditiei"]).reset_index(drop=True)

    # Adăugăm coloana `%`
    data["%"] = 0.6  # Setăm valoarea constantă pentru %

    data["transport"] = ""
    data["STATistica"] = ""

    # Adăugăm totaluri la final
    data = add_totals(data, ["Greutate neta", "Valoare RON", "STATistica"], group_by="Vat/cumparator")
    data = add_excel_formulas(data)

    # Generăm workbook-ul Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    # Scriem anteturile
    for col_num, header in enumerate(Constants.FINAL_HEADERS, 1):
        ws.cell(row=1, column=col_num, value=header)
        ws.cell(row=1, column=col_num).font = Font(bold=True)
        ws.cell(row=1, column=col_num).alignment = Alignment(horizontal="center")

    for i, row in data.iterrows():
        is_falsy = pd.isna(row["Nr. Crt."]) or str(row["Nr. Crt."]).strip() == ""  # Verificăm dacă Nr. Crt. este falsy
        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=i + 2, column=col_num, value=cell_value)
            if is_falsy:
                cell.font = Font(bold=True)  # Aplicăm bold doar dacă Nr. Crt. este falsy

    # Salvăm fișierul Excel
    wb.save(output_path)
    print(f"Excel file saved to {output_path}")


if __name__ == "__main__":
    input_paths = ["pdfs/0912626530_C015000044_ZSM0_001.PDF", "pdfs/9610169997_2007000000_ZSM0_001.PDF"]
    processor = InvoiceProcessor(input_paths)
    processor.process_invoices()
    print(processor.df)

    df = processor.df.rename(columns=Constants.COLUMN_MAPPING)
    current_date = datetime.now().strftime("%d-%m-%Y")
    output_path = f"output/{current_date}-EXP.xlsx"
    generate_excel(df, output_path)
