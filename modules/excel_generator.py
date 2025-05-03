import os
import re
from datetime import datetime, date
from functools import lru_cache

import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.formula.translate import Translator, TranslatorError
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from config import Constants
from functions import convert_to_date, format_nc8_code


class ExcelGenerator:
    """
    Creates and formats an Excel file from a pandas DataFrame.

    This class handles the generation of Excel files from invoice data,
    including formatting, formula generation, and appending to existing files.
    """

    __slots__ = ("data", "percentage")

    def __init__(self, data, percentage=0.6):
        """
        Initialize the ExcelGenerator with invoice data.
        """
        self.data = data
        self.percentage = percentage

    def generate_excel(self, path, existing_excel=None):
        """
        Generate an Excel file from the invoice data.

        Prepare data, add totals, add formulas, then either append to an existing workbook
        or create a new one at 'path'.
        """
        self._prepare_data()

        if existing_excel and os.path.exists(existing_excel):
            self._append_new_invoices_to_workbook(existing_excel, path)
            return

        self._add_totals(["net_weight", "value_ron", "statistic"], group_by="vat_number")
        self._add_excel_formulas()

        wb = Workbook()
        ws = wb.active
        ws.title = "Invoices"

        self._write_headers(ws)
        self._write_rows(ws)

        wb.save(path)

    def _prepare_data(self):
        """
        Sort, clean, and prepare data for Excel output.
        """
        self.data["vat_number"] = self.data["vat_number"].astype(str)
        self.data["invoice_number"] = pd.to_numeric(self.data["invoice_number"], errors="coerce").fillna(0).astype(int)

        self.data["shipment_date"] = self.data["shipment_date"].apply(convert_to_date)
        self.data["nc8_code"] = self.data["nc8_code"].apply(format_nc8_code)

        self.data = self.data.sort_values(by=["vat_number", "shipment_date", "invoice_number"]).reset_index(drop=True)

        self.data.insert(0, "nr_crt", self.data.groupby("vat_number").cumcount() + 1)
        self.data["percentage"] = self.percentage

        for col in ["transport", "statistic"]:
            if col not in self.data.columns:
                self.data[col] = ""
            else:
                self.data[col] = self.data[col].fillna("")

    def _append_new_invoices_to_workbook(self, src_path, dest_path):
        """
        Append non-duplicate invoice records into a copy of the workbook.

        Takes an existing workbook at src_path, adds new records from self.data,
        and saves the result to dest_path, preserving formatting and formulas.
        """
        if not os.path.exists(src_path) or self.data.empty:
            return

        wb = openpyxl.load_workbook(src_path)
        ws = wb.active
        formats = Constants.COLUMN_FORMATS

        header_map = self._map_headers(ws)
        self._remove_total_rows(ws, header_map)
        struct = self._find_data_block(ws, header_map)
        existing = self._collect_existing_keys(ws, struct, header_map)

        df = self.data.dropna(subset=['invoice_number'])
        new_records = [row for _, row in df.iterrows() if
                       (row.vat_number, row.invoice_number, row.nc8_code) not in existing]

        if not new_records:
            return

        for rec in new_records:
            row_idx = self._find_insert_rows(ws, rec, struct, header_map)
            ws.insert_rows(row_idx)

            self._write_record(ws, row_idx, rec, header_map)

            formulas = self._build_formulas_for_row(row_idx, rec)
            self._apply_formulas_and_formatting(ws, row_idx, rec, header_map, formulas, formats)

        self._update_formulas_after_insertion(ws, struct)

        struct = self._find_data_block(ws, header_map)

        self._recompute_nr_crt(ws, struct, header_map)
        self._insert_total_rows(ws, struct, header_map)

        wb.save(dest_path)

    @staticmethod
    def _apply_formulas_and_formatting(ws, row_idx, rec, header_map, formulas, formats):
        """Apply formulas and formatting to a row."""
        for field, col in header_map.items():
            cell = ws.cell(row=row_idx, column=col)

            if field in formulas:
                cell.value = formulas[field]
            else:
                val = rec.get(field)
                if val is not None:
                    cell.value = val

            if field in formats:
                cell.number_format = formats[field]

    @staticmethod
    def _update_formulas_after_insertion(ws, struct):
        """Update cell references in formulas after row insertions."""
        for r in range(struct['data_start'], struct['data_end'] + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=col)
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    cell.value = re.sub(r'([A-Z]+)(\d+)', lambda m: f"{m.group(1)}{r}", formula)

    def _add_excel_formulas(self):
        """
        Add in-cell formulas for computed columns.
        """
        for i, row in self.data.iterrows():
            excel_row = i + 2

            formulas = self._build_formulas_for_row(excel_row, row.to_dict())

            for col, formula in formulas.items():
                self.data.at[i, col] = formula

    def _add_totals(self, columns_to_total, group_by=None):
        """
        Insert total rows after each group or at the end of the DataFrame.
        """
        if not self.data.empty:
            if group_by:
                total_rows = []
                current_row_offset = 1

                grouped = self.data.groupby(group_by, sort=False, as_index=False)

                for name, group in grouped:
                    group_size = len(group)

                    total_row = {col: "" for col in self.data.columns}
                    total_row[group_by] = f"Total {name}"
                    total_row["nr_crt"] = ""

                    for col in columns_to_total:
                        if col in self.data.columns:
                            col_letter = chr(65 + self.data.columns.get_loc(col))
                            start_row = current_row_offset + 1
                            end_row = start_row + group_size - 1
                            total_row[col] = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"

                    current_row_offset += group_size + 1
                    total_rows.append(total_row)

                result_df = pd.DataFrame()
                for (_, group), total_row in zip(grouped, total_rows):
                    result_df = pd.concat([result_df, group.reset_index(drop=True), pd.DataFrame([total_row])],
                                          ignore_index=True)

                self.data = result_df
            else:
                total_row = {col: "" for col in self.data.columns}
                total_row["nr_crt"] = "Total"

                for col in columns_to_total:
                    if col in self.data.columns:
                        col_letter = chr(65 + self.data.columns.get_loc(col))
                        start_row = 2
                        end_row = len(self.data) + 1
                        total_row[col] = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"

                self.data = pd.concat([self.data, pd.DataFrame([total_row])], ignore_index=True)

    @staticmethod
    def _write_headers(ws):
        """
        Write column headers, apply style, and freeze the header row.
        """
        header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        header_font = Font(bold=True, size=Constants.FONT_SIZE, name="Arial")
        header_alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)

        for col_num, header in enumerate(Constants.HEADERS.values(), 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill

        ws.freeze_panes = "A2"

    def _write_rows(self, ws):
        """
        Write DataFrame rows, including formulas, and adjust column widths.
        """
        headers_keys = list(Constants.HEADERS.keys())
        formats = Constants.COLUMN_FORMATS
        regular_font = Font(size=12, name="Arial")
        bold_font = Font(bold=True, size=12, name="Arial")

        col_max_lengths = {col_num: len(header) for col_num, header in enumerate(Constants.HEADERS.values(), 1)}

        for i, row in self.data.iterrows():
            is_total_row = pd.isna(row["nr_crt"]) or str(row["nr_crt"]).strip() == ""
            excel_row = i + 2

            for col_num, cell_value in enumerate(row, 1):
                if cell_value is not None:
                    col_max_lengths[col_num] = max(col_max_lengths[col_num], len(str(cell_value)))

                cell = ws.cell(row=excel_row, column=col_num, value=cell_value)
                header_key = headers_keys[col_num - 1]

                cell.font = bold_font if is_total_row else regular_font

                if header_key in formats:
                    cell.number_format = formats[header_key]

        scaling = Constants.FONT_SIZE / 10 * Constants.SCALING_FACTOR
        for col_num, max_len in col_max_lengths.items():
            col_letter = get_column_letter(col_num)
            adjusted_width = max_len * scaling + 2
            ws.column_dimensions[col_letter].width = adjusted_width

    @staticmethod
    def _map_headers(ws):
        """
        Map header titles to column indices.
        """
        rev = {v: k for k, v in Constants.HEADERS.items()}

        return {rev[cell.value]: i + 1 for i, cell in enumerate(ws[1]) if cell.value in rev}

    @staticmethod
    def _find_data_block(ws, cmap):
        """Identify data range and VAT grouping positions."""
        struct = {'data_start': 2, 'data_end': None, 'blank_row': None, 'vat_groups': {}}
        current_vat = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(not v or (isinstance(v, str) and not v.strip()) for v in row):
                struct['blank_row'], struct['data_end'] = row_idx, row_idx - 1
                break
            vat = row[cmap['vat_number'] - 1]
            num = row[cmap['nr_crt'] - 1]
            if isinstance(vat, str) and vat.startswith('Total '):
                label = vat.split(' ', 1)[1]
                group = struct['vat_groups'].setdefault(label, {})
                group['end'] = row_idx
            elif vat and num and vat != current_vat:
                if current_vat and current_vat in struct['vat_groups'] and struct['vat_groups'][current_vat].get(
                        'end') is None:
                    struct['vat_groups'][current_vat]['end'] = row_idx
                current_vat = vat
                struct['vat_groups'][vat] = {'start': row_idx, 'end': None}

        if struct['blank_row'] is None:
            struct['blank_row'], struct['data_end'] = ws.max_row + 1, ws.max_row

        data_end = struct['data_end']
        for vat, group in struct['vat_groups'].items():
            if group.get('end') is None:
                group['end'] = data_end + 1

        return struct

    @staticmethod
    def _collect_existing_keys(ws, struct, cmap):
        """Collect existing invoice keys to avoid duplicates."""
        keys = set()
        for r in range(struct['data_start'], struct['data_end'] + 1):
            vat = ws.cell(r, cmap['vat_number']).value
            inv = ws.cell(r, cmap['invoice_number']).value
            nc = ws.cell(r, cmap['nc8_code']).value
            if vat and inv and nc and not str(vat).startswith('Total '):
                keys.add((str(vat), int(inv), str(nc)))
        return keys

    @staticmethod
    def _find_insert_rows(ws, rec, struct, cmap):
        """
        Determine the correct row for inserting a new record, updating struct accordingly.
        """
        vat = rec.vat_number

        target_date = ExcelGenerator._normalize_date(rec.shipment_date)
        target = (target_date, rec.invoice_number)

        grp = struct['vat_groups'].get(vat)
        if not grp or not grp.get('end'):
            insert_row = ExcelGenerator._find_new_vat_position(vat, struct)

            ExcelGenerator._update_group_positions(struct, insert_row, vat)

            struct['vat_groups'][vat] = {'start': insert_row, 'end': insert_row + 1}

            struct['data_end'] += 1
            struct['blank_row'] = struct['data_end'] + 1

            return insert_row

        start, end = grp['start'], grp['end']
        insert_row = ExcelGenerator._find_position_in_group(ws, cmap, start, end, target)

        if insert_row is None:
            insert_row = end

        ExcelGenerator._update_group_positions(struct, insert_row, vat)

        struct['vat_groups'][vat]['end'] += 1

        struct['data_end'] += 1
        struct['blank_row'] = struct['data_end'] + 1

        return insert_row

    @staticmethod
    def _normalize_date(date_value):
        """Convert various date formats to a standard date object."""
        if isinstance(date_value, (pd.Timestamp, datetime)):
            return date_value.date()
        elif isinstance(date_value, date):
            return date_value
        return date_value

    @staticmethod
    def _find_new_vat_position(vat, struct):
        """Find the position for a new VAT group."""
        vats = [(v, g['start']) for v, g in struct['vat_groups'].items() if g.get('start') is not None]

        vats.sort(key=lambda x: x[0])

        for v_existing, start_row in vats:
            if v_existing > vat:
                return start_row

        return struct['data_end'] + 1

    @staticmethod
    def _update_group_positions(struct, insert_row, current_vat=None):
        """Update group positions after insertion."""
        for other_v, other_grp in struct['vat_groups'].items():
            if current_vat and other_v == current_vat:
                continue

            if other_grp.get('start') is not None and other_grp['start'] >= insert_row:
                other_grp['start'] += 1

            if other_grp.get('end') is not None and other_grp['end'] >= insert_row:
                other_grp['end'] += 1

    @staticmethod
    def _find_position_in_group(ws, cmap, start, end, target):
        """Find the correct position within a VAT group based on date and invoice number."""
        target_date, target_invoice = target

        for r in range(start, end):
            date_val = ExcelGenerator._get_cell_date(ws, r, cmap['shipment_date'])
            if date_val is None:
                continue

            try:
                inv = int(ws.cell(r, cmap['invoice_number']).value or 0)
            except (ValueError, TypeError):
                continue

            if (date_val, inv) >= (target_date, target_invoice):
                return r

        return None

    @staticmethod
    def _get_cell_date(ws, row, col):
        """Extract and normalize a date value from a cell."""
        raw = ws.cell(row, col).value

        if isinstance(raw, datetime):
            return raw.date()
        elif isinstance(raw, date):
            return raw
        elif isinstance(raw, str):
            try:
                return date.fromisoformat(raw)
            except ValueError:
                pass

        return None

    @staticmethod
    def _write_record(ws, row, rec, cmap):
        """
        Write invoice record into worksheet row.
        """
        regular_font = Font(size=12, name='Arial')

        for field, col in cmap.items():
            val = rec.get(field)

            if field == 'delivery_location' and str(val) == '0':
                alt = next((v for v in rec.delivery_location if v != '0'), None)
                if alt:
                    val = alt

            cell = ws.cell(row=row, column=col, value=val)
            cell.font = regular_font

    @staticmethod
    @lru_cache(maxsize=128)
    def _get_cell_references(row_idx):
        """
        Get Excel cell references for a given row index.
        """
        return {"exchange_rate": f"J{row_idx}", "value_eur": f"G{row_idx}", "value_ron": f"K{row_idx}",
                "net_weight": f"H{row_idx}", "percentage": f"O{row_idx}", "transport": f"P{row_idx}"}

    @staticmethod
    def _build_formulas_for_row(row_idx, row):
        """
        Generate Excel formulas for a given row index and data row.
        """
        if pd.isna(row.get("nr_crt")) or str(row.get("nr_crt")).strip() == "":
            return {}

        cells = ExcelGenerator._get_cell_references(row_idx)

        formulas = {}

        val_eur = row.get("invoice_value_eur", 0)
        val_ron = row.get("value_ron", 0)

        if val_ron == 0 or (val_eur != 0 and val_ron != 0):
            formulas["value_ron"] = f"={cells['value_eur']}*{cells['exchange_rate']}"
        elif val_eur == 0 and val_ron != 0:
            formulas["invoice_value_eur"] = f"={cells['value_ron']}/{cells['exchange_rate']}"

        if pd.notna(row.get("net_weight")) and pd.notna(row.get("exchange_rate")):
            formulas["transport"] = f"=28000*{cells['exchange_rate']}/147000*{cells['net_weight']}"
        else:
            formulas["transport"] = ""

        if pd.notna(val_ron):
            formulas["statistic"] = f"=ROUND({cells['value_ron']}+{cells['percentage']}*{cells['transport']}, 0)"
        else:
            formulas["statistic"] = ""

        return formulas

    @staticmethod
    def _remove_total_rows(ws, header_map):
        """Remove all existing 'Total {vat_number}' rows from the worksheet."""
        vat_col = header_map['vat_number']
        rows_to_delete = []

        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=vat_col).value
            if isinstance(cell_value, str) and cell_value.startswith('Total '):
                rows_to_delete.append(row)

        for row in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row)

    @staticmethod
    def _recompute_nr_crt(ws, struct, header_map):
        """Recompute and rewrite the nr_crt sequence for each VAT block, starting at 1."""
        nr_crt_col = header_map['nr_crt']
        vat_col = header_map['vat_number']

        for vat, group in struct['vat_groups'].items():
            start = group.get('start')
            end = group.get('end')

            if not start or not end:
                continue

            counter = 1

            for row in range(start, end):
                cell_value = ws.cell(row=row, column=vat_col).value
                if isinstance(cell_value, str) and cell_value.startswith('Total '):
                    continue

                ws.cell(row=row, column=nr_crt_col).value = counter
                counter += 1

    @staticmethod
    def _insert_total_rows(ws, struct, header_map):
        """Insert a 'Total {vat_number}' row after each VAT block with SUM formulas."""
        vat_col = header_map['vat_number']

        vat_items = [(vat, grp['start'], grp['end']) for vat, grp in struct['vat_groups'].items() if
                     grp.get('start') and grp.get('end')]
        vat_items.sort(key=lambda x: x[1])

        total_inserted = 0

        for vat, orig_start, orig_end in vat_items:
            new_start = orig_start + total_inserted
            new_end = orig_end + total_inserted

            if total_inserted:
                for row in ws.iter_rows(min_row=new_start, max_row=new_end - 1):
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            try:
                                cell.value = Translator(cell.value, origin=cell.coordinate).translate_formula(
                                    row_delta=total_inserted)
                            except TranslatorError:
                                pass

            ws.insert_rows(new_end)
            total_inserted += 1

            label_cell = ws.cell(row=new_end, column=vat_col, value=f"Total {vat}")
            label_cell.font = Font(bold=True, size=12, name='Arial')

            for field in ('net_weight', 'value_ron', 'statistic'):
                if field in header_map:
                    col_idx = header_map[field]
                    col_letter = get_column_letter(col_idx)
                    sum_formula = f"=SUM({col_letter}{new_start}:{col_letter}{new_end - 1})"
                    sum_cell = ws.cell(row=new_end, column=col_idx, value=sum_formula)
                    sum_cell.font = Font(bold=True, size=12, name='Arial')
                    if field in Constants.COLUMN_FORMATS:
                        sum_cell.number_format = Constants.COLUMN_FORMATS[field]
